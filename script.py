import os
import re
import pandas as pd
from openpyxl.utils import get_column_letter


def read_table(file_path: str) -> pd.DataFrame:
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        return pd.read_csv(file_path)
    elif ext in [".xlsx", ".xls"]:
        return pd.read_excel(file_path)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .csv, .xlsx, or .xls")


def save_table(df: pd.DataFrame, file_path: str) -> None:
    ext = os.path.splitext(file_path)[1].lower()

    if ext == ".csv":
        df.to_csv(file_path, index=False)
    elif ext in [".xlsx", ".xls"]:
        df.to_excel(file_path, index=False)
    else:
        raise ValueError(f"Unsupported file type: {ext}. Use .csv, .xlsx, or .xls")


def normalize_name(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).strip().lower()


def prompt_for_price(item_name: str) -> float:
    print("\nMissing price detected")
    print(f"Item: {item_name}")

    while True:
        val = input("RDU (will be multiplied by 55) > ").strip()
        try:
            return float(val)
        except ValueError:
            print("Please enter a valid number.")


def make_sheet_name(value, used_names: set) -> str:
    if pd.isna(value) or str(value).strip() == "":
        raw = "Blank_Facility"
    else:
        raw = str(value).strip()

    raw = re.sub(r'[:\\/*?\[\]]', "_", raw).strip()
    if not raw:
        raw = "Sheet"

    # Try keeping the start (normal truncation)
    candidate = raw[:31].strip() or "Sheet"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    # Collision after truncation: try keeping the end instead
    # (handles long shared-prefix names like "GC Houston Area Safety Council Baytown")
    tail = raw[-31:].strip()
    # Snap to a word boundary so the name doesn't start mid-word
    space_idx = tail.find(" ")
    if 0 < space_idx < 10:
        tail = tail[space_idx + 1:].strip()
    candidate = tail or "Sheet"
    if candidate not in used_names:
        used_names.add(candidate)
        return candidate

    # Fall back to counter suffix on the start-truncated base
    base = raw[:31].strip() or "Sheet"
    counter = 1
    while True:
        suffix = f"_{counter}"
        name = base[:31 - len(suffix)] + suffix
        if name not in used_names:
            used_names.add(name)
            return name
        counter += 1


def add_price_and_split_by_facility(
    input_file: str,
    input_name_column: str,
    price_file: str,
    price_name_column: str,
    price_column: str,
    facility_column: str,
    output_file: str,
    output_price_column: str = "Price",
) -> None:
    # Read main input file based on extension
    main_df = read_table(input_file)

    # Read or create price file
    if os.path.exists(price_file):
        price_df = read_table(price_file)
    else:
        print(f"\nPrice file '{price_file}' not found. Creating a new one.")
        price_df = pd.DataFrame(columns=[price_name_column, price_column])

    # Validate output format for multi-sheet export
    output_ext = os.path.splitext(output_file)[1].lower()
    if output_ext not in [".xlsx"]:
        raise ValueError(
            "Output file must be .xlsx because the script splits data into multiple sheets."
        )

    if input_name_column not in main_df.columns:
        raise ValueError(
            f"Column '{input_name_column}' not found in input file. "
            f"Available columns: {list(main_df.columns)}"
        )

    if facility_column not in main_df.columns:
        raise ValueError(
            f"Column '{facility_column}' not found in input file. "
            f"Available columns: {list(main_df.columns)}"
        )

    if price_name_column not in price_df.columns:
        raise ValueError(
            f"Column '{price_name_column}' not found in price file. "
            f"Available columns: {list(price_df.columns)}"
        )

    if price_column not in price_df.columns:
        raise ValueError(
            f"Column '{price_column}' not found in price file. "
            f"Available columns: {list(price_df.columns)}"
        )

    price_df = price_df.copy()
    price_df["_match_key"] = price_df[price_name_column].apply(normalize_name)

    duplicates = price_df[
        price_df["_match_key"].ne("") & price_df["_match_key"].duplicated(keep=False)
    ]
    if not duplicates.empty:
        dup_values = duplicates[price_name_column].dropna().astype(str).unique().tolist()
        raise ValueError(
            "Duplicate names found in the price file. Each item should appear only once.\n"
            f"Examples: {dup_values[:10]}"
        )

    price_map = dict(zip(price_df["_match_key"], price_df[price_column]))

    prices = []

    for row_index, name in enumerate(main_df[input_name_column], start=2):
        key = normalize_name(name)

        if key == "":
            print(f"\nRow {row_index}: blank item name, setting price to 0.")
            prices.append(0)
            continue

        if key in price_map:
            prices.append(price_map[key] * 55)
        else:
            print(f"\nRow {row_index} - Unrecognized name: {name}")
            price = prompt_for_price(str(name))
            prices.append(price * 55)

            price_map[key] = price

            new_row = {
                price_name_column: name,
                price_column: price,
                "_match_key": key,
            }
            price_df = pd.concat([price_df, pd.DataFrame([new_row])], ignore_index=True)

            # Save immediately so entered prices are not lost
            save_table(price_df.drop(columns=["_match_key"]), price_file)

    # Create or overwrite price column
    main_df[output_price_column] = prices

    # Sort by facility
    sort_helper = main_df[facility_column].fillna("").astype(str).str.strip().str.lower()
    main_df = main_df.assign(_facility_sort_key=sort_helper).sort_values(
        by="_facility_sort_key",
        kind="stable"
    ).drop(columns=["_facility_sort_key"])

    # Write one sheet per facility
    used_sheet_names = set()

    def _append_sum_formula(ws, df, price_col_name):
        if price_col_name not in df.columns:
            return
        col_idx = list(df.columns).index(price_col_name) + 1  # 1-based
        col_letter = get_column_letter(col_idx)
        last_data_row = len(df) + 1  # +1 for header row
        ws[f"{col_letter}{last_data_row + 1}"] = f"=SUM({col_letter}2:{col_letter}{last_data_row})"

    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        grouped = main_df.groupby(facility_column, dropna=False, sort=False)

        for facility_value, facility_df in grouped:
            sheet_name = make_sheet_name(facility_value, used_sheet_names)
            facility_df.to_excel(writer, sheet_name=sheet_name, index=False)
            _append_sum_formula(writer.sheets[sheet_name], facility_df, output_price_column)

        # Optional combined sheet
        main_df.to_excel(writer, sheet_name="All Data", index=False)
        _append_sum_formula(writer.sheets["All Data"], main_df, output_price_column)

    # Save updated price file one more time at end
    save_table(price_df.drop(columns=["_match_key"]), price_file)

    print("\nFinished.")
    print(f"Output file saved as: {output_file}")
    print(f"Updated price list saved as: {price_file}")
    print(f"Split into sheets by column: {facility_column}")


if __name__ == "__main__":
    import shutil
    from datetime import datetime

    print("=== Price Matching + Facility Sheet Split Script ===\n")

    input_dir = "input"
    price_file = "price_file.xlsx"
    valid_extensions = {".csv", ".xlsx", ".xls"}

    input_name_column = "Study Description"
    price_name_column = "Study Description"
    price_column = "Price"
    facility_column = "Facility"

    if not os.path.isdir(input_dir):
        print(f"Error: '{input_dir}' directory not found.")
        exit(1)

    input_files = [
        f for f in os.listdir(input_dir)
        if os.path.splitext(f)[1].lower() in valid_extensions
    ]

    if not input_files:
        print(f"No valid input files (.csv, .xlsx, .xls) found in '{input_dir}/'.")
        exit(0)

    print(f"Found {len(input_files)} file(s) in '{input_dir}/':\n")
    for f in input_files:
        print(f"  - {f}")
    print()

    # Create a single run directory for this batch
    run_dir = datetime.now().strftime("output_%Y-%m-%d_%H%M%S")
    os.makedirs(run_dir, exist_ok=True)

    for input_filename in input_files:
        input_file = os.path.join(input_dir, input_filename)
        stem = os.path.splitext(input_filename)[0]
        output_file = os.path.join(run_dir, f"{stem}_output.xlsx")

        print(f"\n--- Processing: {input_filename} ---")

        add_price_and_split_by_facility(
            input_file=input_file,
            input_name_column=input_name_column,
            price_file=price_file,
            price_name_column=price_name_column,
            price_column=price_column,
            facility_column=facility_column,
            output_file=output_file,
        )

        shutil.copy2(input_file, os.path.join(run_dir, input_filename))
        os.remove(input_file)
        print(f"Removed input file: {input_file}")

    print(f"\nAll files processed. Run folder: {run_dir}/")